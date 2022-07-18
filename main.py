import openpyxl
# print("hello")

path = "Male-MI-0_Cluster.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row=1, column=1)
#print(cell_obj.value)

maindict = {"ageGroups": {
        "26-30": {"value": 0, "others": {}}, "31-35": {"value": 0, "others": {}}, "36-40": {"value": 0, "others": {}},
        "41-45": {"value": 0, "others": {}}, "46-50": {"value": 0, "others": {}}, "51-55": {"value": 0, "others": {}},
        "56-60": {"value": 0, "others": {}}, "60-65": {"value": 0, "others": {}}, "66-70": {"value": 0, "others": {}}
        , "71-75": {"value": 0, "others": {}}, "76-80": {"value": 0, "others": {}}, "81-85": {"value": 0, "others": {}},
        "86-90": {"value": 0, "others": {}}
    }}
def agecategory(age):
    """
    :param age: int in range of 26 to 90
    :return: agecategory in ageGroups of maindict
    """
    for keys in maindict["ageGroups"]:
        l = keys.split("-")
        print(l)
        if age >= int(l[0]) and age <= int(l[1]):
            return keys

def withvalues(column, totalrow):
    """
    this function is meant to deal with the columns in which there are only certain values which make different
    category.
    ex in ecg we can have 1 or 0 so this creates ecg-1 and ecg0
    :return: void
    """
    for i in range(2, totalrow + 1):
        agecat = agecategory(sheet_obj.cell(row=i, column=1).value)
        if sheet_obj.cell(row=1, column=column).value + "-" + str(sheet_obj.cell(row=i, column=column).value) in \
                maindict["ageGroups"][agecat]["others"]:
            maindict["ageGroups"][agecat]["others"][sheet_obj.cell(row=1, column=column).value + "-" + str(
                sheet_obj.cell(row=i, column=column).value)] += 1
        else:
            maindict["ageGroups"][agecat]["others"][sheet_obj.cell(row=1, column=column).value + "-" + str(
                sheet_obj.cell(row=i, column=column).value)] = 1
    return None


def whichrange(num, rangelist):
    for index, els in enumerate(rangelist):
        l = els.split("-")
        if int(l[0]) <= num and int(l[1]) >= num:
            return index


def withrange(column, totalrow, rangelist):
    for i in range(2, totalrow):
        agecat = agecategory(sheet_obj.cell(row=i, column=1).value)
        # print(whichrange(int(sheet_obj.cell(row=i,column=column).value),rangelist))
        if sheet_obj.cell(row=1, column=column).value + "-" + str(
                whichrange(int(sheet_obj.cell(row=i, column=column).value), rangelist)) in \
                maindict["ageGroups"][agecat]["others"]:
            maindict["ageGroups"][agecat]["others"][sheet_obj.cell(row=1, column=column).value + "-" + str(
                whichrange(sheet_obj.cell(row=i, column=column).value, rangelist))] += 1
        else:
            maindict["ageGroups"][agecat]["others"][sheet_obj.cell(row=1, column=column).value + "-" + str(
                whichrange(sheet_obj.cell(row=i, column=column).value, rangelist))] = 1
    return None


for i in range(2, 650):
    maindict["ageGroups"][agecategory(sheet_obj.cell(row=i, column=1).value)]["value"] += 1
withrange(8, 649, ["124-194", "194-294"])
for i in range(2, 8):
    withvalues(i, 649)
withvalues(9,649)
withvalues(10,649)
print(maindict["ageGroups"])



