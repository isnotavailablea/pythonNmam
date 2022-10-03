import xlsxwriter
import openpyxl
path= "puredata.xlsx"
wb_obj=openpyxl.load_workbook(path)
sheet_obj=wb_obj.active
cell_obj=sheet_obj.cell(row=1,column=1)
workbook=xlsxwriter.Workbook('female.xlsx')
worksheet =workbook.add_worksheet()
worksheet.write(0,0,"Age")
worksheet.write(0,1,"Gender")
worksheet.write(0,2,"ECG")
worksheet.write(0,3,"CKMB")
worksheet.write(0,4,"TROP-I")
worksheet.write(0,5,"Chest_Pain")
worksheet.write(0,6,"Diabetic")
worksheet.write(0,7,"Cholesterol")
worksheet.write(0,8,"PHF /family history")
worksheet.write(0,9,"MI")
ro=1
for i in range(2,897):
    # print(i)
    if(sheet_obj.cell(row=i,column=2).value==0):
        try:
            for j in range(9):
                worksheet.write(ro,j,sheet_obj.cell(row=i,column=j+1).value)
            ro+=1
        except Exception as e:
            print("error ",e)
workbook.close()
