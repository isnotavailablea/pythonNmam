import treeimplemen
#import product
import openpyxl
path="Book1.xlsx"
wb_obj=openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row=1, column=1)

heredict=treeimplemen.scoredict
for i in range(2,1250):
    val=0
    mystr=sheet_obj.cell(row=i,column=3).value+"->"
    for j in range(5,18):
        if sheet_obj.cell(row=i,column=j).value!=0:
            mystr+=sheet_obj.cell(row=1,column=j).value+"->"
            val+=sheet_obj.cell(row=i,column=j).value
            continue
    reqstr=sheet_obj.cell(row=i,column=4).value

    if type(reqstr) is float:
        val+=reqstr
    else:
        reqstr=reqstr[1:]
        l = reqstr.split("/")
        num = float(l[0])
        for i in range(len(l)):
            num /= float(l[i])
        val+=num
    #print(mystr)

    heredict[mystr].append(round(val,3))
j=1
for i in heredict:
    print(f"for path {j} calculated value {heredict[i][0]} and sheet value is {heredict[i][1]}")
    j+=1