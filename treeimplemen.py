import openpyxl

path="finalprod.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row=1, column=1)
maxrow=sheet_obj.max_row
maxcolumn=sheet_obj.max_column

class Head:
    def __init__(self):
        self.next=[]
class Normal:
    def __init__(self,name,value):
        self.name=name
        self.value=value
        self.next=[]

ageBucket=Head()
temp_list=[]
#columns_done=2
for i in range(2,maxrow+1):
    name=sheet_obj.cell(row=i,column=1).value
    value=sheet_obj.cell(row=i,column=2).value
    agenode=Normal(name,value)
    temp_list.append(agenode)
ageBucket.next=temp_list
features=[2,2,3,2,2,2]
index_done=-1



#<---------------------------------------Below is the tree implementation-------------------------------------------------------------------->
queue_current=[]
child_current=[]
temp=[]
for i in ageBucket.next:
    queue_current.append(i)


def insert_node(node, index_done):
    columns_done = 2
    for i in range(index_done + 1):
        columns_done += features[i]
    temp_list = []
    for i in range(features[index_done + 1]):
        name = sheet_obj.cell(row=1, column=columns_done + i + 1).value
        value = sheet_obj.cell(row=2, column=columns_done + i + 1).value
        newnode = Normal(name, value)
        temp_list.append(newnode)
        global child_current
        child_current.append(newnode)
    node.next = temp_list
    print(f"for {node.name} children are:")
    for i in node.next:
        print(i.name, end=" ")
    print("\n")
    return


def addchildren(thelist,index_done):
    if index_done==len(features)-1:
        return
    for i in thelist:
        insert_node(i,index_done)
    global queue_current,child_current
    queue_current=child_current
    child_current=[]
    #index_done+=1
    return addchildren(queue_current,index_done+1)

addchildren(queue_current,-1)





#<-------------------------------------Following is for tree traversal through all possible paths------------------------------------------------->
val=0
def treetraverse(node,pathvalue):
    global val
    if node==[]:
        #print(round(pathvalue,3))
        if round(pathvalue,3)>val:
            val=round(pathvalue,3)
        else:
            pass
        return
    print(node[0].name)
    # if node[0].next==[]:
    #     print(pathvalue+node[0].value)
    #     treetraverse(node[1:],pathvalue)
    #     # return
    treetraverse(node[0].next,pathvalue+node[0].value)
    treetraverse(node[1:],pathvalue)

treetraverse(ageBucket.next,0)
print(val)